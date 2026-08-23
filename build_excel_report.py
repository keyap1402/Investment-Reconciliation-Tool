import pandas as pd

report = pd.read_csv("reconciliation_report.csv")

summary = report["break_type"].value_counts().reset_index()
summary.columns = ["break_type", "count"]

print(summary)

with pd.ExcelWriter("reconciliation_report.xlsx", engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Summary", index=False)
    report.to_excel(writer, sheet_name="Detail", index=False)

# WORKSHEET EDITING
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

wb = load_workbook("reconciliation_report.xlsx")

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for sheet_name in ["Summary", "Detail"]:
    ws = wb[sheet_name]

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 4

wb.save("reconciliation_report.xlsx")
print("Formatting applied.")